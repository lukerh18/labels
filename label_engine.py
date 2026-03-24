"""
label_engine.py  —  NY Price Label Generator core
--------------------------------------------------
Public API:
  build_config(**kwargs)  →  config dict
  generate_workbook(csv_files, cfg)  →  (Workbook, summary)
  render_label_html(item, cfg)  →  HTML string  (used for in-app preview)
  SIZE_PRESETS  →  dict of preset names → (w, h) or None
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import pandas as pd
from datetime import datetime
try:
    import barcode
    from barcode.writer import ImageWriter
    HAS_BARCODE = True
except ImportError:
    HAS_BARCODE = False

import os, tempfile

# ── Constants ─────────────────────────────────────────────────────────────────
BC_DIR = os.path.join(tempfile.gettempdir(), 'lbl_barcodes')
os.makedirs(BC_DIR, exist_ok=True)

today_str  = datetime.now().strftime('%m/%d/%y')
today_date = datetime.now().date()

SIZE_PRESETS = {
    '2" × 1"  — Standard Shelf':   (2.0, 1.0),
    '3" × 1"  — Wide Shelf':        (3.0, 1.0),
    '2" × 1.5" — Tall Shelf':       (2.0, 1.5),
    '4" × 2"  — Large Shelf':       (4.0, 2.0),
    '1.5" × 1" — Small Label':      (1.5, 1.0),
    'Custom':                        None,
}

READ_DTYPES = {'Upc': str, 'Size': str, 'ItemCode': str, 'PLU': str}

# ── Default config ────────────────────────────────────────────────────────────
def build_config(
    # Label size
    label_width_in    = 2.0,
    label_height_in   = 1.0,
    # Orange box
    show_unit_price   = True,
    show_uom          = True,
    show_date         = True,
    # Right column
    show_special_price = False,
    show_multibuy     = True,    # GroupPrice / Quantity  e.g. "2 FOR $5.00"
    show_item_number  = True,
    show_upc          = True,
    show_size         = True,
    show_pack         = True,    # Pack count e.g. "24-pack"
    # Bottom bar
    show_description  = True,
    show_barcode      = True,
    show_snap_badge   = True,    # SNAP / EBT eligible  (Foodstamp=TRUE)
    show_wic_badge    = True,    # WIC eligible  (Wicable=1)
    # Style
    orange_hex        = 'FF8000',
):
    return dict(
        label_width_in=label_width_in, label_height_in=label_height_in,
        show_unit_price=show_unit_price, show_uom=show_uom, show_date=show_date,
        show_special_price=show_special_price, show_multibuy=show_multibuy,
        show_item_number=show_item_number, show_upc=show_upc,
        show_size=show_size, show_pack=show_pack,
        show_description=show_description, show_barcode=show_barcode,
        show_snap_badge=show_snap_badge, show_wic_badge=show_wic_badge,
        orange_hex=orange_hex,
    )

DEFAULT_CONFIG = build_config()

# ── Layout calculator ─────────────────────────────────────────────────────────
BASE_ROW_H = [8, 22, 10, 12, 16, 4]   # pt — sums to 72pt = 1 inch

def get_layout(cfg):
    sw = cfg['label_width_in']  / 2.0
    sh = cfg['label_height_in'] / 1.0
    total_pt = int(cfg['label_height_in'] * 72)
    row_h = [max(4, round(h * total_pt / sum(BASE_ROW_H))) for h in BASE_ROW_H]
    row_h[-1] += total_pt - sum(row_h)

    return dict(
        row_h          = row_h,
        orange_chars   = max(8,  round(11 * sw)),
        price_chars    = max(10, round(16 * sw)),
        gap_chars      = max(2,  round(3  * sw)),
        labels_per_row = max(1, int(10.5 / cfg['label_width_in'])),
        rows_per_label = len(BASE_ROW_H),
        fs_hdr         = max(5,  round(5.5 * sh)),
        fs_unit_price  = max(8,  round(11  * sh)),
        fs_uom         = max(5,  round(6.5 * sh)),
        fs_retail      = max(16, round(22  * sh)),
        fs_was         = max(5,  round(7   * sh)),
        fs_sub         = max(4,  round(5.5 * sh)),
        fs_detail      = max(4,  round(6   * sh)),
        fs_desc        = max(5,  round(6.5 * sh)),
        fs_badge       = max(4,  round(5   * sh)),
    )

# ── Data helpers ──────────────────────────────────────────────────────────────
def get_uom(uom_raw, scale_raw):
    scale = str(scale_raw).strip().upper() == 'TRUE'
    if pd.isna(uom_raw) or str(uom_raw).strip() in ('', 'nan'):
        return 'PER POUND' if scale else 'EACH'
    u = str(uom_raw).strip().upper()
    m = {'LB':'PER POUND','LBS':'PER POUND','OZ':'PER OUNCE','OUNCE':'PER OUNCE',
         'OUNCES':'PER OUNCE','QT':'PER QUART','QUART':'PER QUART',
         'EA':'EACH','EACH':'EACH','CT':'PER 100 CT',
         'FLOZ':'PER FLUID OUNCE','FL OZ':'PER FLUID OUNCE'}
    try: float(u); return 'EACH'
    except ValueError: pass
    return m.get(u, f'PER {u}')

def fmt_upc(upc_raw):
    d = ''.join(filter(str.isdigit, str(upc_raw)))
    if len(d) >= 12:
        d = d.zfill(12)[-12:]
        return f'{d[0]}-{d[1:6]}-{d[6:11]}-{d[11]}'
    return str(upc_raw).strip()

def clean_size(raw):
    s = str(raw).strip() if pd.notna(raw) else ''
    if s in ('', 'nan'): return ''
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else f'{f:g}'
    except ValueError: return s

def make_barcode(upc_raw):
    if not HAS_BARCODE:
        return None
    digits = ''.join(filter(str.isdigit, str(upc_raw)))
    if not digits: return None
    path = os.path.join(BC_DIR, f'{digits}.png')
    if not os.path.exists(path):
        try:
            bc = barcode.get('code128', digits, writer=ImageWriter())
            bc.save(path.replace('.png',''), options={
                'module_width':0.38,'module_height':8.5,'quiet_zone':2.0,
                'font_size':0,'write_text':False,'dpi':300,
                'background':'white','foreground':'black'})
        except Exception: return None
    return path if os.path.exists(path) else None

def is_special_active(item):
    """Returns (is_active: bool, special_price: float|None)."""
    sp = item.get('SpecialPrice','')
    try:
        sp_val = float(sp)
        if sp_val <= 0: return False, None
    except (TypeError, ValueError): return False, None
    sd = item.get('StartDate',''); ed = item.get('EndDate','')
    try:
        start = pd.to_datetime(sd).date() if pd.notna(sd) and str(sd).strip() not in ('','nan') else None
        end   = pd.to_datetime(ed).date() if pd.notna(ed) and str(ed).strip() not in ('','nan') else None
        if start and today_date < start: return False, None
        if end   and today_date > end:   return False, None
    except Exception: pass
    return True, sp_val

def get_multibuy(item):
    """Returns (qty: int, price: float) or (None, None)."""
    try:
        qty = int(float(item.get('Quantity','') or 0))
        gp  = float(item.get('GroupPrice','') or 0)
        if qty > 1 and gp > 0: return qty, gp
    except (TypeError, ValueError): pass
    return None, None

def extract_item_data(item, cfg):
    """Pull all display-ready values from a raw item dict."""
    try:    price = float(item['Price'])
    except: price = 0.0
    pp = item.get('PricePer','')
    try:    price_per = float(pp) if pd.notna(pp) and str(pp).strip() not in ('','nan') else None
    except: price_per = None

    uom      = get_uom(item.get('UnitOfMeasure',''), item.get('Scale','FALSE'))
    desc     = str(item.get('Description','')).upper().strip()
    size     = clean_size(item.get('Size',''))
    upc_raw  = str(item.get('Upc','')).strip()
    upc_fmt  = fmt_upc(upc_raw)
    ic       = str(item.get('ItemCode','')).strip()
    item_code = ic if ic not in ('nan','') else ''

    sd = item.get('StartDate','')
    try:    date_lbl = pd.to_datetime(sd).strftime('%m/%d/%y') if pd.notna(sd) and str(sd).strip() not in ('','nan') else today_str
    except: date_lbl = today_str

    abbr_map = {'PER POUND':'LB','PER OUNCE':'OZ','PER QUART':'QT',
                'PER 100 CT':'CT','PER FLUID OUNCE':'FL OZ'}
    if size:
        if any(c.isalpha() for c in size): size_display = size.upper()
        elif uom and uom != 'EACH':
            a = abbr_map.get(uom,'')
            size_display = f'{size} {a}' if a else size
        else: size_display = size
    else: size_display = ''

    # Pack
    pack_raw = item.get('Pack','')
    try:
        pack_n = int(float(pack_raw))
        pack_str = f'{pack_n}-pack' if pack_n > 1 else ''
    except (TypeError, ValueError): pack_str = ''

    # Badges
    snap = str(item.get('Foodstamp','')).strip().upper() == 'TRUE'
    wic  = str(item.get('Wicable','')).strip() == '1'

    # Special / multi-buy
    use_special, special_val = (False, None)
    if cfg['show_special_price']: use_special, special_val = is_special_active(item)
    mb_qty, mb_price = (None, None)
    if cfg['show_multibuy']: mb_qty, mb_price = get_multibuy(item)

    return dict(
        price=price, price_per=price_per, uom=uom, desc=desc,
        size=size, size_display=size_display, upc_raw=upc_raw, upc_fmt=upc_fmt,
        item_code=item_code, date_lbl=date_lbl, pack_str=pack_str,
        snap=snap, wic=wic, use_special=use_special, special_val=special_val,
        mb_qty=mb_qty, mb_price=mb_price,
    )

# ── HTML preview renderer (used by Streamlit wizard) ─────────────────────────
def render_label_html(item, cfg):
    """Returns an HTML string visually representing the label using real item data."""
    d = extract_item_data(item, cfg)
    orange = f'#{cfg["orange_hex"]}'

    # Right-col sub line: multi-buy > item# > blank
    if d['use_special']:
        sub_right = f'<span style="font-size:9px">WAS ${d["price"]:.2f}'
        if cfg['show_item_number'] and d['item_code']:
            sub_right += f'&nbsp;&nbsp;Item #{d["item_code"]}'
        sub_right += '</span>'
    elif d['mb_qty']:
        sub_right = f'<span style="font-size:9px">{d["mb_qty"]} FOR ${d["mb_price"]:.2f}</span>'
    elif cfg['show_item_number'] and d['item_code']:
        sub_right = f'<span style="font-size:9px">Item #: {d["item_code"]}</span>'
    else:
        sub_right = ''

    # Price display
    if d['use_special']:
        price_html = f'<span style="color:#CC0000;font-size:1.5em;font-weight:bold">${d["special_val"]:.2f}</span>'
        price_hdr  = '<span style="color:#CC0000;font-size:9px">SPECIAL PRICE</span>'
    else:
        price_html = f'<span style="font-size:1.5em;font-weight:bold">${d["price"]:.2f}</span>'
        price_hdr  = '<span style="font-size:9px">RETAIL PRICE</span>'

    # UPC + size + pack detail line
    detail_parts = []
    if cfg['show_upc']:   detail_parts.append(d['upc_fmt'])
    if cfg['show_size'] and d['size_display']:  detail_parts.append(d['size_display'])
    if cfg['show_pack'] and d['pack_str']:       detail_parts.append(d['pack_str'])
    detail_str = '&nbsp;&nbsp;'.join(detail_parts)

    # Badges
    badges = ''
    if cfg['show_snap_badge'] and d['snap']:
        badges += '<span style="background:#2E7D32;color:white;font-size:8px;padding:1px 3px;border-radius:2px;margin-right:2px">SNAP/EBT</span>'
    if cfg['show_wic_badge'] and d['wic']:
        badges += '<span style="background:#1565C0;color:white;font-size:8px;padding:1px 3px;border-radius:2px">WIC</span>'

    # Orange box content
    orange_lines = []
    if cfg['show_unit_price']:
        orange_lines.append('<span style="font-size:9px">UNIT PRICE</span>')
        val = f'${d["price_per"]:.2f}' if d['price_per'] is not None else 'N/A'
        orange_lines.append(f'<b style="font-size:1.1em">{val}</b>')
    if cfg['show_uom']:
        orange_lines.append(f'<span style="font-size:8px">{d["uom"]}</span>')

    html = f"""
    <div style="font-family:Arial,sans-serif;font-size:11px;border:2px solid #444;
                width:100%;display:flex;flex-direction:column;min-height:80px;
                background:white;">
      <!-- Top: orange box + price area -->
      <div style="display:flex;flex:1;">
        <div style="background:{orange};color:white;width:38%;
                    display:flex;flex-direction:column;align-items:center;
                    justify-content:center;text-align:center;padding:4px 2px;
                    gap:1px;border-right:1px solid #aaa;">
          {'<br>'.join(orange_lines) if orange_lines else ''}
        </div>
        <div style="width:62%;display:flex;flex-direction:column;
                    align-items:center;justify-content:center;
                    padding:4px;text-align:center;gap:1px;">
          {price_hdr}
          {price_html}
          {sub_right}
        </div>
      </div>
      <!-- Date / UPC row -->
      <div style="display:flex;border-top:1px solid #ddd;font-size:9px;
                  padding:2px 4px;min-height:14px;">
        <div style="width:38%;color:#555;text-align:center;">
          {''+d['date_lbl'] if cfg['show_date'] else ''}
        </div>
        <div style="width:62%;text-align:center;color:#555;">{detail_str}</div>
      </div>
      <!-- Description / barcode row -->
      <div style="display:flex;border-top:1px solid #555;padding:2px 4px;
                  min-height:16px;align-items:center;">
        <div style="width:60%;font-size:9px;font-weight:bold;">
          {''+d['desc'][:30] if cfg['show_description'] else ''}
          {'&nbsp;' + badges if badges else ''}
        </div>
        <div style="width:40%;text-align:right;font-size:9px;
                    letter-spacing:-1.5px;color:#333;">
          {'▌▌▌▏▌▌▌▌▌▏▌▌▌▌▌' if cfg['show_barcode'] else ''}
        </div>
      </div>
    </div>"""
    return html

# ── Sheet setup ───────────────────────────────────────────────────────────────
def setup_sheet(ws, layout):
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = 1; ws.page_setup.orientation = 'landscape'
    ws.page_margins.left=0.25; ws.page_margins.right=0.25
    ws.page_margins.top=0.35;  ws.page_margins.bottom=0.35
    ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1
    for i in range(layout['labels_per_row']):
        ws.column_dimensions[get_column_letter(i*3+1)].width = layout['orange_chars']
        ws.column_dimensions[get_column_letter(i*3+2)].width = layout['price_chars']
        ws.column_dimensions[get_column_letter(i*3+3)].width = layout['gap_chars']

# ── Single Excel label ────────────────────────────────────────────────────────
def draw_label(ws, r, oc, item, cfg, layout):
    pc  = oc + 1
    MED = Side(style='medium'); THIN = Side(style='thin')
    OFILL = PatternFill('solid', start_color=cfg['orange_hex'], end_color=cfg['orange_hex'])
    d = extract_item_data(item, cfg)

    for i, h in enumerate(layout['row_h']):
        ws.row_dimensions[r+i].height = h

    def oc_set(row, val, sz, bold=False):
        c = ws.cell(row, oc)
        c.value=val; c.fill=OFILL
        c.font=Font(name='Arial',size=sz,bold=bold,color='FFFFFF')
        c.alignment=Alignment(horizontal='center',vertical='center')

    def pc_set(row, val, sz, bold=False, color='000000'):
        c = ws.cell(row, pc)
        c.value=val; c.font=Font(name='Arial',size=sz,bold=bold,color=color)
        c.alignment=Alignment(horizontal='center',vertical='center')

    # Row 0 — headers
    oc_set(r,   'UNIT PRICE' if cfg['show_unit_price'] else '', layout['fs_hdr'])
    if d['use_special']: pc_set(r, 'SPECIAL PRICE', layout['fs_hdr'], color='CC0000')
    else:                pc_set(r, 'RETAIL PRICE',  layout['fs_hdr'])

    # Row 1 — main prices
    oc_set(r+1, (f'${d["price_per"]:.2f}' if d['price_per'] is not None else 'N/A')
                if cfg['show_unit_price'] else '', layout['fs_unit_price'], bold=True)
    if d['use_special']: pc_set(r+1, f'${d["special_val"]:.2f}', layout['fs_retail'], bold=True, color='CC0000')
    else:                pc_set(r+1, f'${d["price"]:.2f}',        layout['fs_retail'], bold=True)

    # Row 2 — UOM | multi-buy / item# / "WAS"
    oc_set(r+2, d['uom'] if cfg['show_uom'] else '', layout['fs_uom'], bold=True)
    if d['use_special']:
        was = f'WAS ${d["price"]:.2f}'
        if cfg['show_item_number'] and d['item_code']: was += f'  Item#:{d["item_code"]}'
        pc_set(r+2, was, layout['fs_was'])
    elif d['mb_qty']:
        pc_set(r+2, f'{d["mb_qty"]} FOR ${d["mb_price"]:.2f}', layout['fs_was'], bold=True)
    else:
        pc_set(r+2, f'Item #: {d["item_code"]}' if (cfg['show_item_number'] and d['item_code']) else '', layout['fs_sub'])

    # Row 3 — date | UPC + size + pack
    c3o = ws.cell(r+3, oc); c3o.value = d['date_lbl'] if cfg['show_date'] else ''
    c3o.font = Font(name='Arial', size=layout['fs_detail'])
    c3o.alignment = Alignment(horizontal='center', vertical='center')

    detail_parts = []
    if cfg['show_upc']:  detail_parts.append(d['upc_fmt'])
    if cfg['show_size'] and d['size_display']:  detail_parts.append(d['size_display'])
    if cfg['show_pack'] and d['pack_str']:       detail_parts.append(d['pack_str'])
    c3p = ws.cell(r+3, pc); c3p.value = '   '.join(filter(None, detail_parts))
    c3p.font = Font(name='Arial', size=layout['fs_detail'])
    c3p.alignment = Alignment(horizontal='center', vertical='center')

    # Row 4 — description + badges | barcode
    badge_parts = []
    if cfg['show_snap_badge'] and d['snap']: badge_parts.append('SNAP/EBT')
    if cfg['show_wic_badge']  and d['wic']:  badge_parts.append('WIC')
    desc_text = d['desc'][:max(10, layout['orange_chars']-2)]
    if badge_parts: desc_text += '  [' + ' · '.join(badge_parts) + ']'

    c4o = ws.cell(r+4, oc); c4o.value = desc_text if cfg['show_description'] else ''
    c4o.font = Font(name='Arial', bold=True, size=layout['fs_desc'])
    c4o.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    if cfg['show_barcode']:
        bc_path = make_barcode(d['upc_raw'])
        if bc_path:
            try:
                img = XLImage(bc_path)
                img.width  = max(60, layout['price_chars'] * 7 - 4)
                img.height = max(14, int(layout['row_h'][4] * 1.33 - 2))
                ws.add_image(img, f'{get_column_letter(pc)}{r+4}')
            except Exception: pass

    # Borders
    ws.cell(r,   oc).border = Border(top=MED, left=MED, right=THIN)
    ws.cell(r+1, oc).border = Border(left=MED, right=THIN)
    ws.cell(r+2, oc).border = Border(left=MED, right=THIN, bottom=THIN)
    ws.cell(r,   pc).border = Border(top=MED, right=MED)
    ws.cell(r+1, pc).border = Border(right=MED)
    ws.cell(r+2, pc).border = Border(right=MED, bottom=THIN)
    ws.cell(r+3, oc).border = Border(left=MED)
    ws.cell(r+3, pc).border = Border(right=MED, bottom=THIN)
    ws.cell(r+4, oc).border = Border(left=MED, bottom=MED)
    ws.cell(r+4, pc).border = Border(right=MED, bottom=MED)

# ── Populate one sheet ────────────────────────────────────────────────────────
def populate_sheet(ws, df, cfg, layout):
    # Filter out inactive items
    if 'Active' in df.columns:
        df = df[df['Active'].astype(str).str.strip().str.upper() != 'FALSE']
    items = df.to_dict('records')
    lpr = layout['labels_per_row']
    col_starts = [1 + i*3 for i in range(lpr)]
    cur_row = 1
    for i in range(0, len(items), lpr):
        for j in range(lpr):
            idx = i + j
            if idx < len(items):
                draw_label(ws, cur_row, col_starts[j], items[idx], cfg, layout)
        cur_row += layout['rows_per_label']
    last_col = col_starts[-1] + 1
    ws.print_area = f'A1:{get_column_letter(last_col)}{cur_row}'
    return len(items)

# ── Public API ────────────────────────────────────────────────────────────────
def generate_workbook(csv_files, cfg=None):
    if cfg is None: cfg = DEFAULT_CONFIG
    layout = get_layout(cfg)
    wb = Workbook(); wb.remove(wb.active)
    summary = []
    for tab_name, file_obj in csv_files:
        tab_name = tab_name[:31]
        try:
            df = pd.read_csv(file_obj, dtype=READ_DTYPES)
            ws = wb.create_sheet(title=tab_name)
            setup_sheet(ws, layout)
            n = populate_sheet(ws, df, cfg, layout)
            summary.append((tab_name, n, None))
        except Exception as e:
            summary.append((tab_name, 0, str(e)))
    return wb, summary
